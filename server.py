#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""电池检测报价查询 - 后端服务器"""
import json
import os
import http.server
import socketserver
import urllib.parse
import openpyxl
import re

PORT = 8765
EXCEL_PATH = r"E:\weixin\报价-2025.9.9.xlsx"
STATIC_DIR = os.path.dirname(os.path.abspath(__file__))

# 管理员密码（简单权限控制）
ADMIN_PASSWORD = "battery2025"

# 标准ID到Excel sheet和行的映射（在启动时建立索引）
standard_index = {}

def build_index():
    """扫描Excel所有sheet，建立标准ID到(sheet_name, start_row, end_row)的映射"""
    global standard_index
    standard_index = {}
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row < 2:
            continue
        
        current_standard = None
        current_start_row = None
        
        for row_idx in range(1, sheet.max_row + 1):
            a1_val = sheet.cell(row=row_idx, column=1).value
            if a1_val and isinstance(a1_val, str) and a1_val.strip():
                # Check if this is a standard title row (contains standard number patterns)
                if re.search(r'(IEC|GB|UL|UN|QC|MT|SJ|ANSI|CCC|ROHS|REACH|CQC)', a1_val, re.IGNORECASE):
                    if current_standard:
                        standard_index[current_standard] = {
                            'sheet': sheet_name,
                            'start_row': current_start_row,
                            'end_row': row_idx - 1
                        }
                    current_standard = a1_val.strip()
                    current_start_row = row_idx
        
        if current_standard:
            standard_index[current_standard] = {
                'sheet': sheet_name,
                'start_row': current_start_row,
                'end_row': sheet.max_row
            }
    
    wb.close()
    print(f"索引已建立，共 {len(standard_index)} 条标准")

def find_standard_in_excel(standard_name):
    """在Excel中查找标准，返回sheet名和行范围"""
    # 精确匹配
    if standard_name in standard_index:
        return standard_index[standard_name]
    
    # 模糊匹配
    for key, val in standard_index.items():
        if standard_name.lower() in key.lower():
            return val
    
    return None

def get_excel_table_data(standard_name):
    """从Excel获取标准的完整表格数据（用于前端展示）"""
    info = find_standard_in_excel(standard_name)
    if not info:
        return None
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb[info['sheet']]
    
    # 读取表头（第2行和第3行）
    headers = []
    for col_idx in range(1, min(sheet.max_column + 1, 13)):
        h1 = sheet.cell(row=info['start_row'] + 1, column=col_idx).value
        h2 = sheet.cell(row=info['start_row'] + 2, column=col_idx).value
        if h1 or h2:
            headers.append({
                'col': col_idx,
                'label1': str(h1) if h1 else '',
                'label2': str(h2) if h2 else ''
            })
    
    # 读取数据行
    rows = []
    for row_idx in range(info['start_row'] + 3, info['end_row'] + 1):
        row_data = []
        has_content = False
        for col_idx in range(1, min(sheet.max_column + 1, 13)):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None:
                has_content = True
            if isinstance(val, float) and val == int(val):
                val = int(val)
            row_data.append({
                'col': col_idx,
                'value': str(val) if val is not None else '',
                'is_number': isinstance(val, (int, float)) and not isinstance(val, bool)
            })
        if has_content:
            rows.append({'excel_row': row_idx, 'cells': row_data})
    
    wb.close()
    
    return {
        'sheet': info['sheet'],
        'title': sheet.cell(row=info['start_row'], column=1).value or '',
        'headers': headers,
        'rows': rows
    }

def get_all_sheets():
    """获取Excel所有sheet名称和基本信息"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheets_info = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheets_info.append({
            'name': sheet_name,
            'rows': sheet.max_row,
            'cols': sheet.max_column
        })
    wb.close()
    return sheets_info

def get_full_sheet_data(sheet_name):
    """获取指定sheet的完整表格数据"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    
    sheet = wb[sheet_name]
    
    # 读取表头（第一行）
    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col_idx).value
        headers.append({
            'col': col_idx,
            'label': str(val) if val is not None else f'列{col_idx}'
        })
    
    # 读取所有数据行
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data = []
        has_content = False
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None:
                has_content = True
            if isinstance(val, float) and val == int(val):
                val = int(val)
            row_data.append({
                'col': col_idx,
                'value': str(val) if val is not None else '',
                'is_number': isinstance(val, (int, float)) and not isinstance(val, bool)
            })
        if has_content:
            rows.append({'excel_row': row_idx, 'cells': row_data})
    
    wb.close()
    
    return {
        'sheet': sheet_name,
        'headers': headers,
        'rows': rows,
        'total_rows': sheet.max_row,
        'total_cols': sheet.max_column
    }


def get_excel_prices(standard_name):
    """从Excel获取标准的价格信息"""
    info = find_standard_in_excel(standard_name)
    if not info:
        return None
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb[info['sheet']]
    
    result = {
        'sheet': info['sheet'],
        'start_row': info['start_row'],
        'title': sheet.cell(row=info['start_row'], column=1).value or '',
        'cell_total_price': 0,
        'battery_total_price': 0,
    }
    
    # 汇总电芯和电池组的总价
    cell_total = 0
    battery_total = 0
    
    for row_idx in range(info['start_row'] + 3, info['end_row'] + 1):
        g_val = sheet.cell(row=row_idx, column=7).value  # 单电芯价格
        h_val = sheet.cell(row=row_idx, column=8).value  # 电池组价格
        
        if g_val and isinstance(g_val, (int, float)):
            cell_total += g_val
        if h_val and isinstance(h_val, (int, float)):
            battery_total += h_val
    
    result['cell_total_price'] = cell_total
    result['battery_total_price'] = battery_total
    
    # 读取威凯报价（I列，可能在第4行）
    for row_idx in range(info['start_row'] + 3, min(info['start_row'] + 8, info['end_row'] + 1)):
        i_val = sheet.cell(row=row_idx, column=9).value
        if i_val and isinstance(i_val, str) and i_val.strip():
            result['weikai_quote'] = i_val.strip()
            break
    
    # 读取证书费
    for row_idx in range(info['start_row'] + 3, min(info['start_row'] + 8, info['end_row'] + 1)):
        j_val = sheet.cell(row=row_idx, column=10).value
        if j_val and isinstance(j_val, str) and j_val.strip():
            result['cert_fee'] = j_val.strip()
            break
    
    # 读取行业参考报价
    for row_idx in range(info['start_row'] + 3, min(info['start_row'] + 8, info['end_row'] + 1)):
        k_val = sheet.cell(row=row_idx, column=11).value
        if k_val and isinstance(k_val, str) and k_val.strip():
            result['industry_price'] = k_val.strip()
            break
    
    wb.close()
    return result

def update_excel_prices(standard_name, cell_price, battery_price):
    """更新Excel和search_index.json中的价格"""
    try:
        cell_price_num = float(cell_price) if cell_price and cell_price.strip() else None
        battery_price_num = float(battery_price) if battery_price and battery_price.strip() else None
    except ValueError:
        return False, "价格格式错误，请输入数字"
    
    if cell_price_num is None and battery_price_num is None:
        return False, "请至少输入一个价格"
    
    cell_price_int = int(cell_price_num) if cell_price_num is not None else None
    battery_price_int = int(battery_price_num) if battery_price_num is not None else None
    
    # 1. 更新 search_index.json
    index_path = os.path.join(STATIC_DIR, 'search_index.json')
    with open(index_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    updated_index = False
    index_data = json.loads(re.search(r'var\s+search_index\s*=\s*(\{.*\})', content, re.DOTALL).group(1))
    
    for key, item in index_data.items():
        if item.get('name') == standard_name or item.get('id') == standard_name:
            if cell_price_int is not None:
                item['cellPrice'] = str(cell_price_int)
            if battery_price_int is not None:
                item['batteryPrice'] = str(battery_price_int)
            updated_index = True
    
    if updated_index:
        new_content = 'var search_index = ' + json.dumps(index_data, ensure_ascii=False, indent=None)
        with open(index_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
    
    # 2. 更新 Excel - 在威凯报价(I列)第一个数据行添加/更新汇总价格
    info = find_standard_in_excel(standard_name)
    if info:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb[info['sheet']]
        
        first_data_row = info['start_row'] + 3
        existing = sheet.cell(row=first_data_row, column=9).value
        
        # 构建更新后的内容，先保留原始内容（排除之前追加的价格行），再追加新价格
        original_lines = []
        if existing and isinstance(existing, str) and existing.strip():
            for line in existing.split('\n'):
                stripped = line.strip()
                # 过滤掉之前通过本系统追加的电芯/电池组价格行
                if re.match(r'^电芯[：:]\s*\d+', stripped) or re.match(r'^电池组/系统[：:]\s*\d+', stripped) or re.match(r'^电池[组系]?[：:]\s*\d+', stripped):
                    continue
                original_lines.append(line)
        
        # 追加新的价格
        if cell_price_int is not None:
            original_lines.append(f"电芯：{cell_price_int}RMB")
        if battery_price_int is not None:
            original_lines.append(f"电池组/系统：{battery_price_int}RMB")
        
        sheet.cell(row=first_data_row, column=9).value = '\n'.join(original_lines)
        
        wb.save(EXCEL_PATH)
        wb.close()
        
        # 重建索引
        build_index()
    
    return True, "价格已保存到Excel和查询系统"

class RequestHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=STATIC_DIR, **kwargs)
    
    def log_message(self, format, *args):
        # 只记录错误和API请求
        if args and len(args) > 1:
            status = args[1] if len(args) > 1 else ''
            if int(status) >= 400:
                print(f"[{self.log_date_time_string()}] {format % args}")
    
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path == '/api/search':
            params = urllib.parse.parse_qs(parsed.query)
            query = params.get('q', [''])[0].strip()
            if not query:
                self.send_json({'error': '缺少查询参数'})
                return
            
            # 从search_index.json搜索
            index_path = os.path.join(STATIC_DIR, 'search_index.json')
            with open(index_path, 'r', encoding='utf-8') as f:
                content = f.read()
                # 提取JSON对象
                match = re.search(r'var\s+search_index\s*=\s*(\{.*\})', content, re.DOTALL)
                if not match:
                    self.send_json({'error': '索引文件格式错误'})
                    return
                search_index = json.loads(match.group(1))
            
            # 搜索逻辑
            if query in search_index:
                item = search_index[query]
            else:
                ql = query.lower()
                if ql in search_index:
                    item = search_index[ql]
                else:
                    matches = []
                    for k, v in search_index.items():
                        if ql in k.lower():
                            matches.append(v)
                    # 去重
                    seen = set()
                    unique = []
                    for m in matches:
                        if m['id'] not in seen:
                            seen.add(m['id'])
                            unique.append(m)
                    if len(unique) == 1:
                        item = unique[0]
                    elif len(unique) > 1:
                        self.send_json({'suggestions': unique[:8]})
                        return
                    else:
                        self.send_json({'not_found': True})
                        return
            
            # 从Excel获取实时价格
            excel_prices = get_excel_prices(item['name'])
            
            result = dict(item)
            if excel_prices:
                result['excel_cell_price'] = excel_prices['cell_total_price']
                result['excel_battery_price'] = excel_prices['battery_total_price']
                result['excel_weikai'] = excel_prices.get('weikai_quote', '')
                result['excel_cert_fee'] = excel_prices.get('cert_fee', '')
                result['excel_industry'] = excel_prices.get('industry_price', '')
                result['excel_sheet'] = excel_prices['sheet']
                result['excel_title'] = excel_prices['title']
            
            self.send_json(result)
        
        elif parsed.path == '/api/table':
            params = urllib.parse.parse_qs(parsed.query)
            standard = params.get('standard', [''])[0]
            if not standard:
                self.send_json({'error': '缺少标准名称'})
                return
            try:
                table_data = get_excel_table_data(standard)
                if table_data:
                    self.send_json({'success': True, 'data': table_data})
                else:
                    self.send_json({'success': False, 'message': '未找到对应标准的表格数据'})
            except Exception as e:
                print(f"获取table失败 (standard={standard}): {e}")
                self.send_json({'success': False, 'message': f'读取表格数据失败: {str(e)}'}, 500)
        
        elif parsed.path == '/api/sheets':
            try:
                sheets = get_all_sheets()
                self.send_json({'success': True, 'data': sheets})
            except Exception as e:
                print(f"获取sheets失败: {e}")
                self.send_json({'success': False, 'message': f'读取Excel失败: {str(e)}'}, 500)
        
        elif parsed.path == '/api/full-sheet':
            params = urllib.parse.parse_qs(parsed.query)
            sheet_name = params.get('sheet', [''])[0]
            if not sheet_name:
                self.send_json({'success': False, 'message': '缺少sheet名称'})
                return
            try:
                data = get_full_sheet_data(sheet_name)
                if data:
                    self.send_json({'success': True, 'data': data})
                else:
                    self.send_json({'success': False, 'message': 'Sheet不存在'})
            except Exception as e:
                print(f"获取full-sheet失败 (sheet={sheet_name}): {e}")
                self.send_json({'success': False, 'message': f'读取Sheet数据失败: {str(e)}'}, 500)
        
        elif parsed.path == '/api/login':
            params = urllib.parse.parse_qs(parsed.query)
            password = params.get('password', [''])[0]
            if password == ADMIN_PASSWORD:
                self.send_json({'success': True, 'message': '登录成功'})
            else:
                self.send_json({'success': False, 'message': '密码错误'})
        
        elif parsed.path == '/api/export':
            # 直接返回Excel文件下载
            try:
                file_size = os.path.getsize(EXCEL_PATH)
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename*=UTF-8\'\'%E6%8A%A5%E4%BB%B7-2025.9.9.xlsx')
                self.send_header('Content-Length', str(file_size))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Cache-Control', 'no-cache')
                self.end_headers()
                with open(EXCEL_PATH, 'rb') as f:
                    while True:
                        chunk = f.read(8192)
                        if not chunk:
                            break
                        self.wfile.write(chunk)
            except Exception as e:
                print(f"导出失败: {e}")
            return
        
        elif parsed.path == '/api/update':
            params = urllib.parse.parse_qs(parsed.query)
            standard = params.get('standard', [''])[0]
            cell_price = params.get('cell', [''])[0]
            battery_price = params.get('battery', [''])[0]
            
            if not standard:
                self.send_json({'success': False, 'message': '缺少标准名称'})
                return
            
            success, message = update_excel_prices(standard, cell_price, battery_price)
            self.send_json({'success': success, 'message': message})
        
        else:
            super().do_GET()
    
    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path == '/api/update':
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    body = self.rfile.read(content_length)
                    data = json.loads(body.decode('utf-8'))
                else:
                    self.send_json({'success': False, 'message': 'Empty request body'}, 400)
                    return
                
                standard = data.get('standard', '')
                cell_price = data.get('cellPrice', '')
                battery_price = data.get('batteryPrice', '')
                
                if not standard:
                    self.send_json({'success': False, 'message': '缺少标准名称'}, 400)
                    return
                
                success, message = update_excel_prices(standard, cell_price, battery_price)
                self.send_json({'success': success, 'message': message})
            except Exception as e:
                self.send_json({'success': False, 'message': str(e)}, 500)
        else:
            self.send_json({'error': 'Not found'}, 404)
    
    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

if __name__ == '__main__':
    build_index()
    print(f"服务器启动: http://localhost:{PORT}")
    with socketserver.ThreadingTCPServer(("", PORT), RequestHandler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n服务器已停止")
