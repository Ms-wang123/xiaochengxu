#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将Excel数据导出为JSON文件，供GitHub Pages静态站点使用"""
import json
import os
import openpyxl
import re

EXCEL_PATH = r"E:\weixin\报价-2025.9.9.xlsx"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

def clean_value(val):
    """清理单元格值"""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        val = int(val)
    return str(val)

def build_standard_index(wb):
    """建立标准ID到sheet和行的索引"""
    index = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row < 2:
            continue
        current_standard = None
        current_start_row = None
        for row_idx in range(1, sheet.max_row + 1):
            a1_val = sheet.cell(row=row_idx, column=1).value
            if a1_val and isinstance(a1_val, str) and a1_val.strip():
                if re.search(r'(IEC|GB|UL|UN|QC|MT|SJ|ANSI|CCC|ROHS|REACH|CQC)', a1_val, re.IGNORECASE):
                    if current_standard:
                        index[current_standard] = {'sheet': sheet_name, 'start_row': current_start_row, 'end_row': row_idx - 1}
                    current_standard = a1_val.strip()
                    current_start_row = row_idx
        if current_standard:
            index[current_standard] = {'sheet': sheet_name, 'start_row': current_start_row, 'end_row': sheet.max_row}
    return index

def export_all():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    standard_index = build_standard_index(wb)
    
    # 1. 导出 sheets 信息
    sheets_info = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheets_info.append({'name': sheet_name, 'rows': sheet.max_row, 'cols': sheet.max_column})
    
    # 2. 导出每个 sheet 的完整数据
    full_sheets = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col_idx).value
            headers.append({'col': col_idx, 'label': clean_value(val) if val is not None else f'列{col_idx}'})
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row_data = []
            has_content = False
            for col_idx in range(1, sheet.max_column + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_content = True
                row_data.append({'col': col_idx, 'value': clean_value(val), 'is_number': isinstance(val, (int, float)) and not isinstance(val, bool)})
            if has_content:
                rows.append({'excel_row': row_idx, 'cells': row_data})
        full_sheets[sheet_name] = {'sheet': sheet_name, 'headers': headers, 'rows': rows, 'total_rows': sheet.max_row, 'total_cols': sheet.max_column}
    
    # 3. 导出每个标准的价格数据
    price_data = {}
    for std_name, info in standard_index.items():
        sheet = wb[info['sheet']]
        # 读取样品数量
        sample_cell = 0
        sample_battery = 0
        cycle_info = ""
        # 读取电芯和电池组价格
        cell_total = 0
        battery_total = 0
        weikai_quote = ""
        cert_fee = ""
        industry_price = ""
        remarks = ""
        
        for row_idx in range(info['start_row'] + 3, info['end_row'] + 1):
            # 样品数量 - 根据Excel结构读取
            b_val = sheet.cell(row=row_idx, column=2).value
            c_val = sheet.cell(row=row_idx, column=3).value
            if b_val and isinstance(b_val, (int, float)):
                sample_cell = int(b_val)
            if c_val and isinstance(c_val, (int, float)):
                sample_battery = int(c_val)
            
            # 周期
            d_val = sheet.cell(row=row_idx, column=4).value
            if d_val and isinstance(d_val, str) and d_val.strip():
                cycle_info = d_val.strip()
            
            # 价格 G=电芯, H=电池组
            g_val = sheet.cell(row=row_idx, column=7).value
            h_val = sheet.cell(row=row_idx, column=8).value
            if g_val and isinstance(g_val, (int, float)):
                cell_total += g_val
            if h_val and isinstance(h_val, (int, float)):
                battery_total += h_val
            
            # 威凯报价 I列
            i_val = sheet.cell(row=row_idx, column=9).value
            if i_val and isinstance(i_val, str) and i_val.strip() and not weikai_quote:
                weikai_quote = i_val.strip()
            
            # 证书费 J列
            j_val = sheet.cell(row=row_idx, column=10).value
            if j_val and isinstance(j_val, str) and j_val.strip() and not cert_fee:
                cert_fee = j_val.strip()
            
            # 行业比价 K列
            k_val = sheet.cell(row=row_idx, column=11).value
            if k_val and isinstance(k_val, str) and k_val.strip() and not industry_price:
                industry_price = k_val.strip()
            
            # 备注 L列
            l_val = sheet.cell(row=row_idx, column=12).value
            if l_val and isinstance(l_val, str) and l_val.strip() and not remarks:
                remarks = l_val.strip()
        
        price_data[std_name] = {
            'title': clean_value(sheet.cell(row=info['start_row'], column=1).value),
            'sample_cell': sample_cell,
            'sample_battery': sample_battery,
            'cycle': cycle_info,
            'cell_price': cell_total,
            'battery_price': battery_total,
            'weikai_quote': weikai_quote,
            'cert_fee': cert_fee,
            'industry_price': industry_price,
            'remarks': remarks,
            'sheet': info['sheet'],
            'start_row': info['start_row'],
            'end_row': info['end_row']
        }
    
    wb.close()
    
    # 写入 data.json
    data = {
        'sheets': sheets_info,
        'full_sheets': full_sheets,
        'prices': price_data
    }
    
    output_path = os.path.join(OUTPUT_DIR, 'data.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    
    print(f"数据已导出到: {output_path}")
    print(f"  - Sheets: {len(sheets_info)} 个")
    print(f"  - 标准价格: {len(price_data)} 条")
    
    # 输出文件大小
    size_mb = os.path.getsize(output_path) / 1024 / 1024
    print(f"  - 文件大小: {size_mb:.2f} MB")

if __name__ == '__main__':
    export_all()
